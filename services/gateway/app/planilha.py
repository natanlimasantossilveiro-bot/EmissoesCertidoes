"""
Adaptado do src/leitor_planilha.py do projeto Certidão Conjunta. Principal
mudança: lê de bytes em memória (upload direto), não de caminho em disco,
já que o Gateway não deve depender de sistema de arquivos local.
"""
import io
from datetime import datetime, date
from openpyxl import load_workbook

COLUNAS_OBRIGATORIAS = ["nome", "tipo", "documento", "data_nascimento"]


def normalizar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def limpar_documento(valor):
    documento = normalizar_texto(valor)
    if documento.endswith(".0"):
        documento = documento[:-2]
    return documento.replace(".", "").replace("-", "").replace("/", "").replace(" ", "")


def normalizar_data(valor):
    if valor is None:
        return ""
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%d/%m/%Y")

    texto = normalizar_texto(valor)
    if not texto:
        return ""

    for formato in ["%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]:
        try:
            return datetime.strptime(texto, formato).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return texto


def ler_planilha_certidoes(conteudo_bytes: bytes):
    workbook = load_workbook(io.BytesIO(conteudo_bytes))
    planilha = workbook.active
    linhas = list(planilha.iter_rows(values_only=True))

    if not linhas:
        raise ValueError("A planilha está vazia.")

    cabecalho = [normalizar_texto(c).lower() for c in linhas[0]]
    for coluna in COLUNAS_OBRIGATORIAS:
        if coluna not in cabecalho:
            raise ValueError(f"Coluna obrigatória não encontrada: {coluna}")

    idx_nome = cabecalho.index("nome")
    idx_tipo = cabecalho.index("tipo")
    idx_documento = cabecalho.index("documento")
    idx_data = cabecalho.index("data_nascimento")

    registros_validos, erros = [], []

    for numero_linha, linha in enumerate(linhas[1:], start=2):
        nome = normalizar_texto(linha[idx_nome])
        tipo = normalizar_texto(linha[idx_tipo]).lower()
        documento = limpar_documento(linha[idx_documento])
        data_nascimento = normalizar_data(linha[idx_data])

        if not nome and not tipo and not documento and not data_nascimento:
            continue

        erros_linha = []
        if not nome:
            erros_linha.append("nome é obrigatório")
        if tipo not in ["pf", "pj"]:
            erros_linha.append("tipo deve ser 'pf' ou 'pj'")
        if not documento:
            erros_linha.append("documento é obrigatório")
        if tipo == "pf" and len(documento) != 11:
            erros_linha.append("CPF deve conter 11 dígitos")
        if tipo == "pj" and len(documento) != 14:
            erros_linha.append("CNPJ deve conter 14 dígitos")
        if tipo == "pf" and not data_nascimento:
            erros_linha.append("data_nascimento é obrigatória para PF")

        if erros_linha:
            erros.append({"linha": numero_linha, "erros": erros_linha})
            continue

        registros_validos.append({
            "linha": numero_linha,
            "nome": nome,
            "tipo": tipo,
            "documento": documento,
            "data_nascimento": data_nascimento,
        })

    workbook.close()
    return registros_validos, erros
